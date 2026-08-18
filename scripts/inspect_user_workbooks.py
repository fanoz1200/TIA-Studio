"""Read-only workbook inventory for user-supplied .xlsx files.

This script never evaluates formulas or executes workbook content. It only reports
sheet metadata and a short preview of non-empty displayed cells.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def cell_preview(value: object) -> str:
    text = str(value).replace("\n", " ").strip()
    return text[:140] + ("…" if len(text) > 140 else "")


def inspect(path_text: str) -> None:
    path = Path(path_text)
    print(f"=== {path.name} ===")
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    for worksheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            values = [cell_preview(value) for value in row if value not in (None, "")]
            if values:
                rows.append(values)
            if len(rows) == 4:
                break
        print(f"[Sheet] {worksheet.title} | rows={worksheet.max_row} | cols={worksheet.max_column}")
        for index, values in enumerate(rows, start=1):
            print(f"  {index}: " + " | ".join(values))


if __name__ == "__main__":
    for workbook_path in sys.argv[1:]:
        inspect(workbook_path)
