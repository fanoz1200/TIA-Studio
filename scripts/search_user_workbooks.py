"""Search visible cell text in user-supplied .xlsx files without evaluating formulas."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit("Usage: search_user_workbooks.py <needle> <workbook> [...]")

    needle = sys.argv[1].casefold()
    for path_text in sys.argv[2:]:
        path = Path(path_text)
        print(f"=== {path.name} ===")
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        found = 0
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=False):
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).replace("\n", " ").strip()
                    if needle in text.casefold():
                        print(f"{worksheet.title}!{cell.coordinate}: {text[:300]}")
                        found += 1
                        if found == 30:
                            break
                if found == 30:
                    break
            if found == 30:
                break
        if not found:
            print("(لا توجد نتيجة مطابقة)")


if __name__ == "__main__":
    main()
